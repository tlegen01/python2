import openpyxl
from openpyxl import Workbook

class MyCalculator:
    @staticmethod  # дакоратор который делает метод в классе статическим(без параметра селф и инициализации)
    def summ(val1: float, val2: float):
        return val1 + val2

    def subtraction(val1: float, val2: float):
        return val1 - val2

    def multiplication(val1: float, val2: float):
        return val1 * val2

    def division(val1: float, val2: float):
        if val2 != 0:
            return val1 / val2
        else:
            print("vy vveli 0")
            return 0


MyCalculator.division(15, 17)
print(MyCalculator.division(15, 17))

##############################################################################################################

class Generator:
    def __init__(self, name: str, index: str, value: str):
        self.name = name
        self.index = index
        self.value = value
    def get_name(self):
        return f"{self.name} {self.index} {self.value}"

ran = []
for i in range(1, 1001):
    ran.append(Generator(name=f"_{i}", index=f"{i}", value=f"_A{i}").get_name())
print(ran)

###############################################################################################################
from openpyxl.utils import get_column_letter

filename = "_svodnye_tablicy.xlsx"
filedir = "home_work"
workbook = openpyxl.load_workbook(filedir + "/" + filename)
worksheet = workbook.active
max_row = worksheet.max_row
max_column = worksheet.max_column

tovar = []


class Produrty:
    def __init__(self, tovar: str, group: str, postavshik: str, data_postavki: float, region: str, prodazhi: float,
                 sbyt: float, pribyl: str):
        self.tovar = tovar
        self.group = group
        self.postavshik = postavshik
        self.data_postavki = data_postavki
        self.region = region
        self.prodazhi = prodazhi
        self.sbyt = sbyt
        self.pribyl = pribyl

    def zarobotok(self):
        return f"{self.tovar},{self.prodazhi},{self.pribyl}"
    def pribyl_dir(self):
        for i in self.pribyl:
            if i == "Да":
                return self.pribyl



obshee_masiv1 = []
for row in range(1, max_row + 1):
    tovar_add = []
    for col in range(1, max_column + 1):
        obj = worksheet.cell(row=row, column=col).value
        # print(obj)
        tovar_add.append(obj)
    # print(tovar_add)
    naimenovanie = Produrty(
        tovar=tovar_add[0],
        group=tovar_add[1],
        postavshik=tovar_add[2],
        data_postavki=tovar_add[3],
        region=tovar_add[4],
        prodazhi=tovar_add[5],
        sbyt=tovar_add[6],
        pribyl=tovar_add[7]
    )
    obshee = naimenovanie.zarobotok()
    pribl = naimenovanie.pribyl_dir()
    # print(obshee)
    obshee_masiv = obshee.split(",")
    # print(obshee_masiv)

    obshee_masiv1.append(obshee_masiv)
# print(obshee_masiv1)
kniga1 = Workbook()
stranica1 = kniga1.active
# for row in range(0, len(obshee_masiv1)):
#     for col in range(0, len(obshee_masiv1[row])):
#         stranica1.cell(row=col + 1, column=row + 1).value = obshee_masiv1[row][col]

for row in range(0, len(obshee_masiv1)):
    for col in range(0, len(obshee_masiv1[row])):
        # print(stroka_final[row])
        # col_letter = get_column_letter(col)
        stranica1[f"{get_column_letter(col + 1)}{row + 1}"] = obshee_masiv1[row][col]
print(obshee_masiv1)

kniga1.save("home_work/oop.xlsx")