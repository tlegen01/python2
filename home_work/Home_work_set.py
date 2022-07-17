import openpyxl
workbook = openpyxl.load_workbook("data.xlsx")
worksheet = workbook.active
max_row = worksheet.max_row + 1
max_column = worksheet.max_column + 1

# external_array = []
# for row in range(1, max_row + 1):
#     internal_array = []
#     for col in range(1, max_column + 1):
#         value = worksheet.cell(row=row, column=col).value
#         if value is None:
#             value = ""
#         internal_array.append(value)
#     external_array.append(internal_array)
# print(external_array)
get1 = []
get2 = []
get3 = []

for i in range(1, max_column +1):
    get1.append(worksheet.cell(2, i).value)
print(get1)

for i in range(1, max_column +1):
    get2.append(worksheet.cell(4, i).value)
print(get2)

for i in range(1, max_column + 1):
    get3.append(worksheet.cell(6, i).value)
print(get3)

set1 = set(get1)
set1.remove(None)
print(set1)
set2 = set(get2)
set2.remove(None)
print(set2)
set3 = set(get3)
set3.remove(None)
print(set3)

# set4 = set2.intersection(set1, set3)
# print(set4)
#
# set5 = set2.difference(set1, set3)
# print(set5)
#
# set6 = set3.union(set1)
# print(set6)

