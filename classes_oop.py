import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

filename = "users.xlsx"
filedir = "temp"
workbook = openpyxl.load_workbook(filedir + "/" + filename)
worksheet = workbook.active
max_row = worksheet.max_row
max_column = worksheet.max_column

workers = []
class Worker:  # определения (создания базовых параметров) класса
    def __init__(self, first_name: str, second_name: str, patronymic: str, position, category):  # магический метод для инициализации (создания) класса
        self.value_first_name = first_name
        self.value_second_name = second_name
        self.value_patronymic = patronymic
        self.value_position = position
        self.value_category = category

    def print_full_name(self):
        print(f'{self.value_second_name} {self.value_first_name}')

    def get_full_name(self):
        return f'{self.value_second_name} {self.value_first_name}'


for row in range (1, max_row +1):
    worker = []
    for column in range(1, 5+1):
        obj = worksheet.cell(row=row, column=column) #метод (функция в классе) который получает объект по координатам
        # print(value)
        value = obj.value #свойство (переменная в классе) которая получает значение ячейки
        if not value:
            value = ""
        worker.append(value)
    worker_obj = Worker(
        first_name=worker[1],
        second_name=worker[0],
        patronymic=worker[2],
        position=worker[3],
        category=worker[4])
    # worker_obj.print_full_name()
    full_name = worker_obj.get_full_name()
    # print(full_name)
    # workers.append(worker)
    workers.append(worker_obj)
workers[4].print_full_name()
    # print(worker_obj)
# print(workers[1].value_first_name)
# workbook = Workbook()



# print(workbook)
# print(type(workbook))

class CustomWorkerClass1:  # определения (создания базовых параметров) класса
    def __init__(self):  # магический метод для инициализации (создания) класса
        self.VALUE_CONSTANT = 123  # свойство переменная внутри класса


worker = CustomWorkerClass1()  # создание экземпляра класса CustomWorkerClass
# print(worker)
# print(type(worker))
# print(worker.VALUE_CONSTANT)  # получение свойства экземпляра класса CustomWorkerClass

external_value_variable = 1

class CustomWorkerClass2:  # определения (создания базовых параметров) класса
    def __init__(self, external_value_variable2):  # магический метод для инициализации (создания) класса
        self.value_veriable = external_value_variable2  # свойство переменная внутри класса

worker = CustomWorkerClass2(156)
# print(worker)
# print(type(worker))
# print(worker.value_veriable)