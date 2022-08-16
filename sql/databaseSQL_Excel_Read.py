import psycopg2
import openpyxl

workbook = openpyxl.load_workbook("/home/tlegen/Документы/GitHub/python2/home_work/_svodnye_tablicy.xlsx")
worksheet = workbook.active
max_row = worksheet.max_row
max_column = worksheet.max_column
global_rows = []
for row in range(2, max_row + 1):
    local_row = []
    for col in range(1, max_column + 1):
        local_row.append(worksheet.cell(row=row, column=col).value)
    global_rows.append(local_row)
print(global_rows)


# query_string = f"""
# INSERT INTO public.products (tovar, gruppa, postavshick, date_post, region, prodashi, sbit, pribil)
# VALUES ('1111111111111', '111111', '111111', '2022-06-06', '111111', '600', '500', 'true')"""
# cursor.execute(query_string)
# conn.commit()

class Record:
    def __init__(self, row: list):
        self.tovar = str(row[0])
        self.gruppa = str(row[1])
        self.postavshick = str(row[2])
        #################################
        date_post = str(row[3]).split(" ")[0]
        year = date_post.split('-')[0]
        month = int(date_post.split('-')[1])
        if int(month) < 10:
            month = f"0{month}"
        day = int(date_post.split('-')[2])
        if int(day) < 10:
            month = f"0{day}"
        self.date_post = str(f"{year}-{month}-{day}")
        ###############################################
        self.region = str(row[4])
        try:
            self.prodashi = int(row[5])
        except Exception as error:
            print(error)
            self.prodashi = 0
        try:
            self.sbit = int(row[6])
        except Exception as error:
            # print(error)
            self.sbit = 0
        self.pribil = bool(row[7])
        pribil = str(row[7])
        if pribil.lower() == "да":
            self.pribil = True
        else:
            self.pribil = False

conn = psycopg2.connect(dbname='example', user='postgres', password='Qwerty123', host='localhost')
cursor = conn.cursor()

for row in global_rows:
    try:
        obj = Record(row=row)
        query_string = f"""
        INSERT INTO public.products (tovar, gruppa, postavshick, date_post, region, prodashi, sbit, pribil)
        VALUES ('{obj.tovar}', '{obj.gruppa}', '{obj.postavshick}', '{obj.date_post}', '{obj.region}', '{obj.prodashi}', 
        '{obj.sbit}', '{obj.pribil}')"""
        cursor.execute(query_string)
        conn.commit()
    except Exception as error:
        print(error)
        print(row)
cursor.close()
conn.close()
