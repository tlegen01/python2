import psycopg2
import openpyxl
from openpyxl.utils import get_column_letter

conn = psycopg2.connect(dbname='example', user='postgres', password='Qwerty123', host='localhost')
cursor = conn.cursor()
cursor.execute("""SELECT * FROM public.products
ORDER BY id ASC""")
records = cursor.fetchall()
print(records)
print(type(records))
cursor.close()
conn.close()

#########################################
book = openpyxl.Workbook()
sheet = book.active
titles = ["Загаловок 1", "Загаловок 2", "Загаловок 3", "Загаловок 4", "Загаловок 5"]
index_col = 1
for title in titles:
    sheet[f"{get_column_letter(index_col)}1"] = str(title)
    index_col += 1
row_index = 2
for row in records:
    col_index = 1
    for col in row:
        sheet[f"{get_column_letter(col_index)}{row_index}"] = str(col)
        col_index += 1
    row_index += 1
book.save('products.xlsx')